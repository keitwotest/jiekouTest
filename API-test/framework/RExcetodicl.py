#coding:UTF-8
"""
--------------------------------------
   File Name：  test_radio
   Description :
   Author :    admin
   Date：     2018/9/13
--------------------------------------
"""
#coding:UTF-8
"""
--------------------------------------
   File Name：  RExcetodicl.py
   Description :
   Author :    admin
   Date：     2018/9/13
--------------------------------------
"""
from openpyxl import load_workbook
from framework.logger import Logger


logger = Logger(logger="RWExcel").getlog()


class RExcetodicl():
    """读取Excel数据"""
    def RExcetodicl(self,path_ex):
        wb = load_workbook(path_ex)
        # 获取所有表名
        sheet_names = wb.sheetnames  # 得到工作簿的所有工作表名 结果： ['Sheet1', 'Sheet2', 'Sheet3']
        # 根据表名打开sheet表
        sheet1 = wb[sheet_names[0]]  # 打开第一个 sheet 工作表
        # 获取C列的所有数据
        datas=[]
        for b, e, c, a, d,f,g,h in zip(sheet1["B"][1:], sheet1["E"][1:], sheet1["C"][1:], sheet1["A"][1:],sheet1["D"][1:], sheet1["F"][1:],sheet1["G"][1:],sheet1["H"][1:]):
            data={
                "url":c.value,
                "param": f.value,
                "method": d.value,
                "headers": e.value,
                "testname": b.value,
                "expect": g.value,
                "case_class":h.value,
                "path_ex":path_ex,
                "row":a.value,
            }
            if data["case_class"]=="K":
                datas.append(data)

        try:
            wb.save(path_ex)
            #logger.info('测试数据保存成功！！！')
        except Exception as e:
            logger.error(e)
            logger.error('保存失败，可能Excel文件未关闭，请关闭Excel文件后重新测试')

        return  {
                "message": "ok",
                "status": "1000",
                "data": datas
            }






