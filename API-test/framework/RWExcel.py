#coding:UTF-8
"""
--------------------------------------
   File Name：  RWExcel.py
   Description :
   Author :    admin
   Date：     2018/9/13
--------------------------------------
"""

from openpyxl  import load_workbook
from framework.Method import Merhod
from framework.logger import Logger
from framework.WriteData import WriteData
logger = Logger(logger="RWExcel").getlog()

class RWExcel():
    def RW_Excel(self,path_ex):
        passs=[]
        fails=[]
        errors=[]

        wb = load_workbook(path_ex)
        # 获取所有表名
        sheet_names = wb.sheetnames  # 得到工作簿的所有工作表名 结果： ['Sheet1', 'Sheet2', 'Sheet3']
        # 根据表名打开sheet表
        sheet1 = wb[sheet_names[0]]  # 打开第一个 sheet 工作表

        # 获取C列的所有数据
        list_sheet1_column_E = []  # 请求参数
        list_sheet1_column_B = []  # 请求地址
        list_sheet1_column_C = []  # 请求方式
        list_sheet1_column_A = []  # 请求名称项
        list_sheet1_column_F = []  # 期望关键字

        for b, e, c, a, f in zip(sheet1["B"], sheet1["E"], sheet1["C"], sheet1["A"], sheet1["F"]):
            list_sheet1_column_A.append(a.value)
            list_sheet1_column_B.append(b.value)
            list_sheet1_column_C.append(c.value)
            list_sheet1_column_E.append(e.value)
            list_sheet1_column_F.append(f.value)
        row=2
        for url, method, str_param, testname, expect in zip(list_sheet1_column_B[1:], list_sheet1_column_C[1:],
                                                            list_sheet1_column_E[1:], list_sheet1_column_A[1:],
                                                            list_sheet1_column_F[1:]):
            #logger.info(str(url), str(method), str(str_param), str(testname))
            PM=(Merhod.merhod(self,url, method, str_param, testname))

            try:
                if WriteData.writedata(self, PM, expect, sheet1, row, testname)=='fail':# 读写excel
                    fails.append('fail')
                else:#WriteData.writedata(self, PM, expect, sheet1, row, testname)=='pass':# 读写excel
                    passs.append('pass')


                #WriteData.writedata(self, PM, expect, sheet1, row, testname)


            except Exception as e:
                logger.error(e)
                logger.error('《' + str(testname) + '》项，写“error”数据！！！')
                sheet1.cell(row=row, column=7, value=e)  # 响应结果
                sheet1.cell(row=row, column=8, value='error')  # 请求时间
                sheet1.cell(row=row, column=9, value='error')  # 状态码
                sheet1.cell(row=row, column=10, value='error')  # 判断通过
                errors.append('error')

            row += 1
        try:
            wb.save(path_ex)
            #logger.info('测试数据保存成功！！！')

        except Exception as e:
            logger.error(e)
            logger.error('保存失败，可能Excel文件未关闭，请关闭Excel文件后重新测试')
        return (
            {
                "pass":len(passs),
                "fail":len(fails),
                "error":len(errors),
            }
        )








