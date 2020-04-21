#coding:UTF-8
"""
--------------------------------------
   File Name：  test_radio
   Description :
   Author :    admin
   Date：     2018/9/13
--------------------------------------
"""
#coding:UTF-8
"""
--------------------------------------
   File Name：  WExcel.py
   Description :
   Author :    admin
   Date：     2018/9/13
--------------------------------------
"""
from openpyxl  import load_workbook
from framework.logger import Logger
from framework.WriteData import WriteData

logger = Logger(logger="WExcel").getlog()

class WExcel():

    def WExcel(self,path_ex,PM,expect,row,testname):
        wb = load_workbook(path_ex)
        # 获取所有表名
        sheet_names = wb.sheetnames  # 得到工作簿的所有工作表名 结果： ['Sheet1', 'Sheet2', 'Sheet3']
        # 根据表名打开sheet表
        sheet1 = wb[sheet_names[0]]  # 打开第一个 sheet 工作表

        WriteData.writedata(self,PM,expect,sheet1,row,testname)

        try:
            wb.save(path_ex)
            #logger.info('测试数据保存成功！！！')
        except Exception as e:
            logger.error(e)
            logger.error('保存失败，可能Excel文件未关闭，请关闭Excel文件后重新测试')
